from openpyxl import load_workbook
from io import BytesIO

def read_xlsx(file_bytes: bytes) -> str:
    wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    texts = []

    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            row_text = " | ".join(
                [str(cell).strip() for cell in row if cell is not None]
            )
            if row_text:
                texts.append(row_text)

    return "\n".join(texts)
